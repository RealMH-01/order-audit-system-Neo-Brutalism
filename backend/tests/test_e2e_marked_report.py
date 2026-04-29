import hashlib
import io
import json
import sys
import zipfile
from pathlib import Path

from docx import Document
from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.services.evidence_locator.cell_index import build_cell_index
from app.services.evidence_locator.resolver import resolve_issue_locations
from app.services.marked_workbook_generator import generate_marked_copies
from app.services.report_generator import ReportGeneratorService


TIMESTAMP = "20260429-2100"
DETAIL_FILENAME = f"审核详情版-H-{TIMESTAMP}.xlsx"


def _rgb(cell) -> str:
    return str(cell.fill.fgColor.rgb)[-6:]


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _save(workbook: Workbook, path: Path) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path


def _single_invoice(path: Path, *, existing_fill=False, existing_comment=False, filename_stem="商业发票") -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["E9"] = "Invoice No."
    sheet["F9"] = "HR-EXP250401"
    sheet["C18"] = "Unit Price"
    sheet["C19"] = 1.85
    if existing_fill:
        sheet["F9"].fill = PatternFill(fill_type="solid", fgColor="FF00FF00")
    if existing_comment:
        sheet["F9"].comment = Comment("original reviewer note", "legacy")
    return _save(workbook, path / f"{filename_stem}.xlsx")


def _multi_sheet_invoice(path: Path) -> Path:
    workbook = Workbook()
    first = workbook.active
    first.title = "SheetA"
    first["E9"] = "Invoice No."
    first["F9"] = "HR-EXP250401"
    second = workbook.create_sheet("SheetB")
    second["E9"] = "Invoice No."
    second["F9"] = "HR-EXP250402"
    return _save(workbook, path / "多sheet发票.xlsx")


def _merged_invoice(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["B3"] = "Invoice No."
    sheet.merge_cells("C3:D4")
    sheet["C3"] = "MERGED-INV-001"
    return _save(workbook, path / "合并单元格.xlsx")


def _formula_invoice(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["B19"] = 10
    sheet["C18"] = "Unit Price"
    sheet["C19"] = "=B19*2"
    return _save(workbook, path / "公式发票.xlsx")


def _duplicate_invoice(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["E9"] = "Invoice No."
    sheet["F9"] = "DUP-001"
    sheet["E23"] = "Invoice No."
    sheet["F23"] = "DUP-001"
    return _save(workbook, path / "重复值.xlsx")


def _write_misc_fixtures(path: Path) -> dict[str, Path]:
    path.mkdir(parents=True, exist_ok=True)
    xls = path / "旧版发票.xls"
    xls.write_bytes(b"not a modern xlsx workbook")

    xlsm = path / "宏发票.xlsm"
    workbook = Workbook()
    workbook.active["A1"] = "macro-enabled placeholder"
    workbook.save(xlsm)

    pdf = path / "扫描发票.pdf"
    pdf.write_bytes(b"%PDF-1.4\n1 0 obj<<>>endobj\ntrailer<<>>\n%%EOF")

    docx = path / "说明文件.docx"
    document = Document()
    document.add_paragraph("supporting document")
    document.save(docx)
    return {"xls": xls, "xlsm": xlsm, "pdf": pdf, "docx": docx}


def _uploaded(file_id: str, path: Path, *, filename: str | None = None, extension: str | None = None) -> dict:
    name = filename or path.name
    return {
        "id": file_id,
        "filename": name,
        "extension": extension or Path(name).suffix.lower().lstrip("."),
        "original_xlsx_path": str(path) if (extension or Path(name).suffix.lower().lstrip(".")) == "xlsx" else "",
    }


def _issue(
    issue_id: str,
    *,
    file_id: str,
    file_name: str,
    field_name: str,
    your_value: str,
    source_value: str,
    hint: str | None,
    level: str = "RED",
    document_type: str = "invoice",
    finding: str = "可定位单元格发现错误",
) -> dict:
    return {
        "id": issue_id,
        "level": level,
        "field_name": field_name,
        "finding": finding,
        "suggestion": f"改为 {source_value}" if source_value else "请人工复核",
        "your_value": your_value,
        "source_value": source_value,
        "document_type": document_type,
        "file_id": file_id,
        "file_name": file_name,
        "location_hints": [hint] if hint else [],
        "confidence": 0.98,
    }


def _resolve(issue: dict, workbook_path: Path, uploaded_file: dict) -> dict:
    file_id = uploaded_file["id"]
    index = build_cell_index(str(workbook_path), file_id, uploaded_file["filename"], issue.get("document_type"))
    locations, status, code, reason = resolve_issue_locations(
        issue,
        {file_id: index},
        {file_id: uploaded_file},
    )
    resolved = dict(issue)
    resolved["locations"] = locations
    resolved["mark_status"] = status
    resolved["mark_reason_code"] = code
    resolved["mark_reason"] = reason
    if issue.get("candidate_locations"):
        resolved["candidate_locations"] = issue["candidate_locations"]
    return resolved


def _unsupported_issue(issue_id: str, uploaded_file: dict, *, extension: str, level: str = "RED") -> dict:
    base = _issue(
        issue_id,
        file_id=uploaded_file["id"],
        file_name=uploaded_file["filename"],
        field_name="invoice_no",
        your_value="HR-EXP250401",
        source_value="HR-EXP250400",
        hint="Sheet1!F9",
        level=level,
        finding=f"{extension} 文件中的问题仅进入详情版",
    )
    locations, status, code, reason = resolve_issue_locations(base, {uploaded_file["id"]: []}, {uploaded_file["id"]: uploaded_file})
    return {**base, "locations": locations, "mark_status": status, "mark_reason_code": code, "mark_reason": reason}


def _advisory_issue(issue_id: str, level: str) -> dict:
    base = {
        "id": issue_id,
        "level": level,
        "field_name": "supporting_documents",
        "finding": "建议补充外部资料并人工确认流程",
        "suggestion": "补充资料",
        "your_value": "",
        "source_value": "",
        "document_type": "invoice",
        "file_id": "xlsx-1",
        "location_hints": [],
    }
    locations, status, code, reason = resolve_issue_locations(
        base,
        {"xlsx-1": [{"file_id": "xlsx-1", "file_name": "商业发票.xlsx", "document_type": "invoice", "sheet": "Sheet1", "cell": "F9", "value_str": "HR-EXP250401"}]},
        {"xlsx-1": {"id": "xlsx-1", "filename": "商业发票.xlsx", "extension": "xlsx"}},
    )
    return {**base, "locations": locations, "mark_status": status, "mark_reason_code": code, "mark_reason": reason}


def _summary(issues: list[dict]) -> dict:
    counts = {"red": 0, "yellow": 0, "blue": 0}
    for issue in issues:
        level = str(issue.get("level", "")).upper()
        if level == "RED":
            counts["red"] += 1
        elif level == "BLUE":
            counts["blue"] += 1
        else:
            counts["yellow"] += 1
    return counts


def _zip(issues: list[dict], uploaded_files: list[dict]) -> io.BytesIO:
    context = {"uploaded_files": uploaded_files}
    filenames = {
        "marked": f"审核标记版-H-{TIMESTAMP}.xlsx",
        "detailed": DETAIL_FILENAME,
        "zip": f"审核报告-H-{TIMESTAMP}.zip",
    }
    return ReportGeneratorService().generate_report_zip(
        "task-e2e",
        {"confidence": 0.91, "summary": _summary(issues), "issues": issues},
        context,
        filenames,
    )


def _marked_names(zip_file: zipfile.ZipFile) -> list[str]:
    return [name for name in zip_file.namelist() if name.startswith("标记版/") and name.endswith(".xlsx")]


def _load_marked(zip_file: zipfile.ZipFile, name_part: str):
    name = next(name for name in _marked_names(zip_file) if name_part in name)
    return load_workbook(io.BytesIO(zip_file.read(name)), data_only=False)


def _detail_rows(zip_file: zipfile.ZipFile) -> dict[str, tuple]:
    workbook = load_workbook(io.BytesIO(zip_file.read(DETAIL_FILENAME)), data_only=False)
    sheet = workbook.worksheets[0]
    return {row[0]: row for row in sheet.iter_rows(min_row=2, values_only=True)}


def _detail_headers(zip_file: zipfile.ZipFile) -> list[str]:
    workbook = load_workbook(io.BytesIO(zip_file.read(DETAIL_FILENAME)), data_only=False)
    assert workbook.sheetnames[:2] == ["问题明细", "审核摘要"]
    return [cell.value for cell in workbook.worksheets[0][1]]


def test_fixture_setup_prepares_all_required_boundary_files(tmp_path):
    fixtures = tmp_path / "fixtures"
    paths = {
        "single": _single_invoice(fixtures),
        "multi_sheet": _multi_sheet_invoice(fixtures),
        "merged": _merged_invoice(fixtures),
        "formula": _formula_invoice(fixtures),
        "duplicates": _duplicate_invoice(fixtures),
        "existing_fill": _single_invoice(fixtures, existing_fill=True, filename_stem="已有背景色"),
        "existing_comment": _single_invoice(fixtures, existing_comment=True, filename_stem="已有批注"),
        **_write_misc_fixtures(fixtures),
    }

    assert {path.suffix.lower() for path in paths.values()} >= {".xlsx", ".xls", ".xlsm", ".pdf", ".docx"}
    assert all(path.exists() and path.stat().st_size > 0 for path in paths.values())


def test_e1_single_xlsx_marks_invoice_no_and_unit_price_and_zip_structure(tmp_path):
    source = _single_invoice(tmp_path)
    upload = _uploaded("xlsx-1", source)
    issues = [
        _resolve(
            _issue("E1-F9", file_id="xlsx-1", file_name=source.name, field_name="invoice_no", your_value="HR-EXP250401", source_value="HR-EXP250400", hint="Sheet1!F9", finding="发票号错值"),
            source,
            upload,
        ),
        _resolve(
            _issue("E1-C19", file_id="xlsx-1", file_name=source.name, field_name="unit_price", your_value="1.85", source_value="1.83", hint="Sheet1!C19", finding="单价错值"),
            source,
            upload,
        ),
    ]

    with zipfile.ZipFile(_zip(issues, [upload])) as zip_file:
        assert zip_file.namelist() == [
            DETAIL_FILENAME,
            f"标记版/审核标记版-{source.stem}-{TIMESTAMP}.xlsx",
            "任务信息.txt",
            "manifest.json",
        ]
        marked = _load_marked(zip_file, source.stem)
        assert _rgb(marked["Sheet1"]["F9"]) == "FF6B6B"
        assert _rgb(marked["Sheet1"]["C19"]) == "FF6B6B"
        assert marked["Sheet1"]["F9"].comment is not None
        assert marked["Sheet1"]["C19"].comment is not None
        assert _detail_headers(zip_file)[:12] == [
            "序号",
            "级别",
            "字段",
            "问题说明",
            "建议",
            "原文件名",
            "原表位置",
            "定位置信度",
            "标记状态",
            "未标记原因",
            "文档类型",
            "文件 ID",
        ]
        detail = _detail_rows(zip_file)
        assert detail["E1-F9"][6] == "Sheet1!F9"
        assert detail["E1-F9"][8] == "marked"
        assert detail["E1-F9"][9] is None
        assert detail["E1-C19"][6] == "Sheet1!C19"
        assert detail["E1-C19"][8] == "marked"
        manifest = json.loads(zip_file.read("manifest.json").decode("utf-8"))
        assert manifest["schema_version"] == "1.0"
        assert manifest["summary"]["by_mark_status"]["marked"] == 2


def test_e2_multiple_xlsx_same_original_name_get_distinct_marked_filenames(tmp_path):
    file_a = _single_invoice(tmp_path / "a", filename_stem="CI")
    file_b = _single_invoice(tmp_path / "b", filename_stem="CI")
    upload_a = _uploaded("xlsx-a", file_a, filename="商业发票.xlsx")
    upload_b = _uploaded("xlsx-b", file_b, filename="商业发票.xlsx")
    issues = [
        _resolve(_issue("E2-A", file_id="xlsx-a", file_name="商业发票.xlsx", field_name="invoice_no", your_value="HR-EXP250401", source_value="HR-EXP250400", hint="Sheet1!F9"), file_a, upload_a),
        _resolve(_issue("E2-B", file_id="xlsx-b", file_name="商业发票.xlsx", field_name="invoice_no", your_value="HR-EXP250401", source_value="HR-EXP250400", hint="Sheet1!F9"), file_b, upload_b),
    ]

    with zipfile.ZipFile(_zip(issues, [upload_a, upload_b])) as zip_file:
        marked = _marked_names(zip_file)
        assert len(marked) == 2
        assert len(set(marked)) == 2


def test_e3_multi_sheet_marks_only_target_sheet(tmp_path):
    source = _multi_sheet_invoice(tmp_path)
    upload = _uploaded("xlsx-1", source)
    issue = _resolve(
        _issue("E3", file_id="xlsx-1", file_name=source.name, field_name="invoice_no", your_value="HR-EXP250402", source_value="HR-EXP250400", hint="SheetB!F9"),
        source,
        upload,
    )

    with zipfile.ZipFile(_zip([issue], [upload])) as zip_file:
        marked = _load_marked(zip_file, source.stem)
        assert _rgb(marked["SheetB"]["F9"]) == "FF6B6B"
        assert _rgb(marked["SheetA"]["F9"]) != "FF6B6B"


def test_e4_merged_cell_hint_writes_to_anchor(tmp_path):
    source = _merged_invoice(tmp_path)
    upload = _uploaded("xlsx-1", source)
    issue = _resolve(
        _issue("E4", file_id="xlsx-1", file_name=source.name, field_name="invoice_no", your_value="MERGED-INV-001", source_value="MERGED-INV-000", hint="Sheet1!D4"),
        source,
        upload,
    )

    with zipfile.ZipFile(_zip([issue], [upload])) as zip_file:
        marked = _load_marked(zip_file, source.stem)
        assert issue["locations"][0]["cell"] == "C3"
        assert _rgb(marked["Sheet1"]["C3"]) == "FF6B6B"


def test_e5_formula_cell_fill_changes_but_formula_is_preserved(tmp_path):
    source = _formula_invoice(tmp_path)
    upload = _uploaded("xlsx-1", source)
    issue = _resolve(
        _issue("E5", file_id="xlsx-1", file_name=source.name, field_name="unit_price", your_value="=B19*2", source_value="18", hint="Sheet1!C19", finding="公式单元格结果疑似错误"),
        source,
        upload,
    )

    with zipfile.ZipFile(_zip([issue], [upload])) as zip_file:
        marked = _load_marked(zip_file, source.stem)
        assert marked["Sheet1"]["C19"].value == "=B19*2"
        assert _rgb(marked["Sheet1"]["C19"]) == "FF6B6B"


def test_e6_duplicate_value_is_multiple_candidates_and_not_colored(tmp_path):
    source = _duplicate_invoice(tmp_path)
    upload = _uploaded("xlsx-1", source)
    issue = _resolve(
        _issue("E6", file_id="xlsx-1", file_name=source.name, field_name="invoice_no", your_value="DUP-001", source_value="DUP-000", hint=None),
        source,
        upload,
    )

    marked_paths, summary, updated = generate_marked_copies([issue], {"uploaded_files": [upload]}, tmp_path / "out", TIMESTAMP)
    assert issue["mark_status"] == "multiple_candidates"
    assert issue["locations"] == []
    assert marked_paths == []
    assert summary == []
    assert updated[0]["mark_status"] == "multiple_candidates"


def test_e7_existing_fill_is_overridden_only_in_copy_and_original_hash_is_unchanged(tmp_path):
    source = _single_invoice(tmp_path, existing_fill=True, filename_stem="已有背景色")
    original_hash = _sha256(source)
    upload = _uploaded("xlsx-1", source)
    issue = _resolve(
        _issue("E7", file_id="xlsx-1", file_name=source.name, field_name="invoice_no", your_value="HR-EXP250401", source_value="HR-EXP250400", hint="Sheet1!F9"),
        source,
        upload,
    )

    with zipfile.ZipFile(_zip([issue], [upload])) as zip_file:
        marked = _load_marked(zip_file, source.stem)
        assert _rgb(marked["Sheet1"]["F9"]) == "FF6B6B"
    assert _sha256(source) == original_hash
    original = load_workbook(source)
    assert _rgb(original["Sheet1"]["F9"]) == "00FF00"


def test_e8_existing_comment_copy_contains_audit_comment_and_original_is_unchanged(tmp_path):
    source = _single_invoice(tmp_path, existing_comment=True, filename_stem="已有批注")
    upload = _uploaded("xlsx-1", source)
    issue = _resolve(
        _issue("E8", file_id="xlsx-1", file_name=source.name, field_name="invoice_no", your_value="HR-EXP250401", source_value="HR-EXP250400", hint="Sheet1!F9", finding="覆盖或合并批注测试"),
        source,
        upload,
    )

    with zipfile.ZipFile(_zip([issue], [upload])) as zip_file:
        marked = _load_marked(zip_file, source.stem)
        assert marked["Sheet1"]["F9"].comment is not None
        assert "覆盖或合并批注测试" in marked["Sheet1"]["F9"].comment.text
    original = load_workbook(source)
    assert original["Sheet1"]["F9"].comment.text == "original reviewer note"


def test_e9_e10_xls_and_xlsm_do_not_generate_marked_versions_but_keep_detail_reasons(tmp_path):
    misc = _write_misc_fixtures(tmp_path)
    upload_xls = _uploaded("xls-1", misc["xls"], extension="xls")
    upload_xlsm = _uploaded("xlsm-1", misc["xlsm"], extension="xlsm")
    issues = [
        _unsupported_issue("E9", upload_xls, extension="xls"),
        _unsupported_issue("E10", upload_xlsm, extension="xlsm"),
    ]

    with zipfile.ZipFile(_zip(issues, [upload_xls, upload_xlsm])) as zip_file:
        assert _marked_names(zip_file) == []
        detail = _detail_rows(zip_file)
        assert detail["E9"][8] == "unsupported_file_type"
        assert detail["E10"][8] == "unsupported_file_type"
        assert detail["E9"][9]
        assert detail["E10"][9]


def test_e11_pdf_and_xlsx_mixed_only_generates_xlsx_marked_version(tmp_path):
    source = _single_invoice(tmp_path)
    misc = _write_misc_fixtures(tmp_path)
    upload_xlsx = _uploaded("xlsx-1", source)
    upload_pdf = _uploaded("pdf-1", misc["pdf"], extension="pdf")
    issues = [
        _resolve(_issue("E11-XLSX", file_id="xlsx-1", file_name=source.name, field_name="invoice_no", your_value="HR-EXP250401", source_value="HR-EXP250400", hint="Sheet1!F9"), source, upload_xlsx),
        _unsupported_issue("E11-PDF", upload_pdf, extension="pdf", level="YELLOW"),
    ]

    with zipfile.ZipFile(_zip(issues, [upload_xlsx, upload_pdf])) as zip_file:
        assert len(_marked_names(zip_file)) == 1
        detail = _detail_rows(zip_file)
        assert detail["E11-XLSX"][8] == "marked"
        assert detail["E11-PDF"][8] == "unsupported_file_type"


def test_e12_yellow_and_blue_advisory_issues_are_not_applicable_and_not_marked(tmp_path):
    source = _single_invoice(tmp_path)
    upload = _uploaded("xlsx-1", source)
    issues = [_advisory_issue("E12-Y", "YELLOW"), _advisory_issue("E12-B", "BLUE")]

    with zipfile.ZipFile(_zip(issues, [upload])) as zip_file:
        assert _marked_names(zip_file) == []
        detail = _detail_rows(zip_file)
        assert detail["E12-Y"][8] == "not_applicable"
        assert detail["E12-B"][8] == "not_applicable"


def test_e13_single_file_mark_failure_keeps_other_marked_file(tmp_path):
    good = _single_invoice(tmp_path / "good", filename_stem="good")
    bad = tmp_path / "bad" / "bad.xlsx"
    bad.parent.mkdir(parents=True)
    bad.write_text("not an xlsx", encoding="utf-8")
    upload_good = _uploaded("good", good)
    upload_bad = _uploaded("bad", bad)
    good_issue = _resolve(
        _issue("E13-good", file_id="good", file_name=good.name, field_name="invoice_no", your_value="HR-EXP250401", source_value="HR-EXP250400", hint="Sheet1!F9"),
        good,
        upload_good,
    )
    bad_issue = {
        **_issue("E13-bad", file_id="bad", file_name=bad.name, field_name="invoice_no", your_value="HR-EXP250401", source_value="HR-EXP250400", hint="Sheet1!F9"),
        "locations": [{"file_id": "bad", "file_name": bad.name, "sheet": "Sheet1", "cell": "F9", "confidence": 0.95}],
        "mark_status": "marked",
        "mark_reason_code": "MARKED",
        "mark_reason": "test setup",
    }

    with zipfile.ZipFile(_zip([good_issue, bad_issue], [upload_good, upload_bad])) as zip_file:
        assert len(_marked_names(zip_file)) == 1
        detail = _detail_rows(zip_file)
        assert detail["E13-good"][8] == "marked"
        assert detail["E13-bad"][8] == "mark_failed"


def test_e14_all_pdf_upload_has_no_empty_marked_directory_and_keeps_metadata(tmp_path):
    misc = _write_misc_fixtures(tmp_path)
    upload_pdf = _uploaded("pdf-1", misc["pdf"], extension="pdf")
    issue = _unsupported_issue("E14", upload_pdf, extension="pdf", level="YELLOW")

    with zipfile.ZipFile(_zip([issue], [upload_pdf])) as zip_file:
        names = zip_file.namelist()
        assert DETAIL_FILENAME in names
        assert "任务信息.txt" in names
        assert "manifest.json" in names
        assert not any(name.startswith("标记版/") for name in names)


def test_e15_chinese_filename_is_preserved_in_zip_namelist(tmp_path):
    source = _single_invoice(tmp_path, filename_stem="中文商业发票")
    upload = _uploaded("xlsx-1", source)
    issue = _resolve(
        _issue("E15", file_id="xlsx-1", file_name=source.name, field_name="invoice_no", your_value="HR-EXP250401", source_value="HR-EXP250400", hint="Sheet1!F9"),
        source,
        upload,
    )

    with zipfile.ZipFile(_zip([issue], [upload])) as zip_file:
        assert any(name == f"标记版/审核标记版-中文商业发票-{TIMESTAMP}.xlsx" for name in zip_file.namelist())
