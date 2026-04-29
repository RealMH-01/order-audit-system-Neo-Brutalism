import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook, load_workbook

from app.services.marked_workbook_generator import generate_marked_copies, merge_levels
from app.services.report_generator import ReportGeneratorService


def _save_workbook(path: Path, *, merge: bool = False, formula: bool = False) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = 1
    sheet["A2"] = 2
    sheet["C3"] = "=SUM(A1:A2)" if formula else "plain"
    sheet["F9"] = "HR-EXP250401"
    sheet["C19"] = "USD 900"
    if merge:
        sheet.merge_cells("A1:B2")
    workbook.save(path)
    return path


def _uploaded(file_id: str, path: Path, filename: str = "CI-test.xlsx", extension: str = "xlsx") -> dict:
    return {
        file_id: {
            "id": file_id,
            "filename": filename,
            "extension": extension,
            "original_xlsx_path": str(path),
        }
    }


def _issue(
    issue_id: str,
    cell: str,
    *,
    file_id: str = "file-1",
    file_name: str = "CI-test.xlsx",
    level: str = "RED",
    finding: str = "发票号错值 HR-EXP250401",
    suggestion: str = "改为 HR-EXP250400",
) -> dict:
    return {
        "id": issue_id,
        "level": level,
        "field_name": "invoice_no",
        "finding": finding,
        "suggestion": suggestion,
        "mark_status": "marked",
        "mark_reason_code": "MARKED",
        "locations": [
            {
                "file_id": file_id,
                "file_name": file_name,
                "sheet": "Sheet1",
                "cell": cell,
                "confidence": 0.95,
            }
        ],
    }


def _rgb(cell) -> str:
    return str(cell.fill.fgColor.rgb)[-6:]


def test_merge_levels_prefers_highest_severity():
    assert merge_levels(["blue", "YELLOW"]) == "YELLOW"
    assert merge_levels(["BLUE", "RED", "YELLOW"]) == "RED"


def test_single_issue_single_cell_sets_fill_and_comment(tmp_path):
    source = _save_workbook(tmp_path / "source.xlsx")

    marked_paths, summary, updated_issues = generate_marked_copies(
        [_issue("R-01", "F9")],
        _uploaded("file-1", source),
        tmp_path / "out",
        "20260429-1901",
    )

    assert len(marked_paths) == 1
    assert summary[0]["status"] == "generated"
    assert updated_issues[0]["mark_status"] == "marked"

    workbook = load_workbook(marked_paths[0], data_only=False)
    cell = workbook["Sheet1"]["F9"]
    assert _rgb(cell) == "FF6B6B"
    assert cell.comment is not None
    assert cell.comment.author == "审核系统"
    assert "发票号错值" in cell.comment.text
    assert "建议: 改为 HR-EXP250400" in cell.comment.text


def test_multiple_issues_same_cell_uses_highest_color_and_combined_comment(tmp_path):
    source = _save_workbook(tmp_path / "source.xlsx")
    issues = [
        _issue("B-01", "F9", level="BLUE", finding="提示问题"),
        _issue("R-01", "F9", level="RED", finding="严重问题"),
    ]

    marked_paths, _, _ = generate_marked_copies(
        issues,
        _uploaded("file-1", source),
        tmp_path / "out",
        "20260429-1902",
    )

    workbook = load_workbook(marked_paths[0], data_only=False)
    cell = workbook["Sheet1"]["F9"]
    assert _rgb(cell) == "FF6B6B"
    assert "本单元格命中 2 个问题" in cell.comment.text
    assert cell.comment.text.index("[RED] 严重问题") < cell.comment.text.index("[BLUE] 提示问题")


def test_merged_cell_location_writes_fill_to_anchor_cell(tmp_path):
    source = _save_workbook(tmp_path / "source.xlsx", merge=True)

    marked_paths, _, _ = generate_marked_copies(
        [_issue("R-01", "B2")],
        _uploaded("file-1", source),
        tmp_path / "out",
        "20260429-1903",
    )

    workbook = load_workbook(marked_paths[0], data_only=False)
    sheet = workbook["Sheet1"]
    assert _rgb(sheet["A1"]) == "FF6B6B"
    assert sheet["A1"].comment is not None


def test_formula_cell_keeps_formula_value_while_marked(tmp_path):
    source = _save_workbook(tmp_path / "source.xlsx", formula=True)

    marked_paths, _, _ = generate_marked_copies(
        [_issue("R-01", "C3", finding="公式结果疑似错误")],
        _uploaded("file-1", source),
        tmp_path / "out",
        "20260429-1904",
    )

    workbook = load_workbook(marked_paths[0], data_only=False)
    cell = workbook["Sheet1"]["C3"]
    assert cell.value == "=SUM(A1:A2)"
    assert _rgb(cell) == "FF6B6B"
    assert "公式结果疑似错误" in cell.comment.text


def test_load_failure_marks_issue_copy_failed_and_other_files_still_generate(tmp_path):
    good_source = _save_workbook(tmp_path / "good.xlsx")
    bad_source = tmp_path / "bad.xlsx"
    bad_source.write_text("not an xlsx", encoding="utf-8")
    issues = [
        _issue("R-good", "F9", file_id="good", file_name="good.xlsx", finding="好文件问题"),
        _issue("R-bad", "F9", file_id="bad", file_name="bad.xlsx", finding="坏文件问题"),
    ]

    marked_paths, summary, updated_issues = generate_marked_copies(
        issues,
        {
            **_uploaded("good", good_source, "good.xlsx"),
            **_uploaded("bad", bad_source, "bad.xlsx"),
        },
        tmp_path / "out",
        "20260429-1905",
    )

    assert len(marked_paths) == 1
    assert {item["file_id"]: item["status"] for item in summary} == {"good": "generated", "bad": "failed"}
    assert updated_issues[0]["mark_status"] == "marked"
    assert updated_issues[1]["mark_status"] == "mark_failed"
    assert updated_issues[1]["mark_reason_code"] == "WRITE_FAILED"
    assert issues[1]["mark_status"] == "marked"


def test_pdf_issue_is_skipped_in_summary_and_keeps_unsupported_status(tmp_path):
    source = _save_workbook(tmp_path / "source.xlsx")
    issues = [
        _issue("R-01", "F9"),
        {
            "id": "PDF-01",
            "level": "YELLOW",
            "finding": "PDF 资料缺失",
            "suggestion": "补充资料",
            "file_id": "pdf-1",
            "mark_status": "unsupported_file_type",
            "mark_reason_code": "FILE_TYPE_NOT_EXCEL",
            "mark_reason": "非 Excel 文件",
            "locations": [],
        },
    ]

    marked_paths, summary, updated_issues = generate_marked_copies(
        issues,
        {
            **_uploaded("file-1", source),
            "pdf-1": {"id": "pdf-1", "filename": "scan.pdf", "extension": "pdf"},
        },
        tmp_path / "out",
        "20260429-1906",
    )

    assert len(marked_paths) == 1
    assert {item["file_id"]: item["status"] for item in summary} == {
        "file-1": "generated",
        "pdf-1": "skipped",
    }
    assert updated_issues[1]["mark_status"] == "unsupported_file_type"


def test_generate_marked_copies_returns_updated_issues_for_detail_report(tmp_path):
    source = _save_workbook(tmp_path / "ci-source.xlsx")
    issues = [_issue("R-01", "F9", finding="F9 错值描述")]

    marked_paths, _, updated_issues = generate_marked_copies(
        issues,
        _uploaded("file-1", source, "ci-source.xlsx"),
        tmp_path / "out",
        "20260429-1906",
    )
    result = {"summary": {"red": 1, "yellow": 0, "blue": 0}, "issues": updated_issues}
    detail = ReportGeneratorService().generate_detail_report("task-1", result)
    workbook = load_workbook(detail)
    headers = [cell.value for cell in workbook["问题明细"][1]]

    assert detail.getvalue()
    assert workbook.sheetnames[:2] == ["问题明细", "审核摘要"]
    assert headers[:12] == [
        "序号",
        "级别",
        "字段",
        "问题说明",
        "建议",
        "原文件名",
        "原表位置",
        "定位置信度",
        "标记状态",
        "未标记原因",
        "文档类型",
        "文件 ID",
    ]
    row = next(workbook["问题明细"].iter_rows(min_row=2, values_only=True))
    assert row[6] == "Sheet1!F9"
    assert row[8] == "marked"
    assert row[9] is None
    assert workbook["审核摘要"]["A1"].value == "任务 ID"
    assert updated_issues is not issues
    assert len(marked_paths) == 1
