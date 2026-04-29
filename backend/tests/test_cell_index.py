import sys
from pathlib import Path

from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from app.services.evidence_locator.cell_index import build_cell_index, normalize_merged_cell


def _save_workbook(tmp_path, workbook, filename="CI-test.xlsx"):
    path = tmp_path / filename
    workbook.save(path)
    return path


def _by_cell(index):
    return {(item["sheet"], item["cell"]): item for item in index}


def test_build_cell_index_single_sheet_simple_file(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = "Invoice No."
    sheet["B1"] = "HR-EXP250401"
    path = _save_workbook(tmp_path, workbook)

    index = build_cell_index(str(path), "file-1", path.name)
    records = _by_cell(index)

    assert records[("Sheet1", "A1")]["document_type"] == "invoice"
    assert records[("Sheet1", "B1")]["value_str"] == "HR-EXP250401"
    assert records[("Sheet1", "B1")]["left_label"] == "Invoice No."
    assert records[("Sheet1", "B1")]["row_context"] == "Invoice No.|HR-EXP250401"


def test_build_cell_index_multi_sheet_file(tmp_path):
    workbook = Workbook()
    workbook.active.title = "Invoice"
    workbook.active["A1"] = "Invoice No."
    workbook.active["B1"] = "INV-001"
    packing = workbook.create_sheet("Packing")
    packing["A1"] = "Cartons"
    packing["B1"] = 20
    path = _save_workbook(tmp_path, workbook, "PL-test.xlsx")

    records = _by_cell(build_cell_index(str(path), "file-1", path.name))

    assert ("Invoice", "B1") in records
    assert ("Packing", "B1") in records
    assert records[("Packing", "B1")]["document_type"] == "packing_list"


def test_build_cell_index_merged_cells_only_outputs_anchor_and_normalizes_children(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.merge_cells("A1:B2")
    sheet["A1"] = "Merged Header"
    path = _save_workbook(tmp_path, workbook)

    index = build_cell_index(str(path), "file-1", path.name)
    records = _by_cell(index)

    assert ("Sheet1", "A1") in records
    assert ("Sheet1", "B1") not in records
    assert ("Sheet1", "A2") not in records
    assert ("Sheet1", "B2") not in records
    assert records[("Sheet1", "A1")]["merged_range"] == "A1:B2"
    assert records[("Sheet1", "A1")]["is_merge_anchor"] is True
    assert normalize_merged_cell("Sheet1", "B2", ["A1:B2"]) == "A1"
    assert normalize_merged_cell("Sheet1", "C3", ["A1:B2"]) == "C3"


def test_build_cell_index_formula_preserves_formula_string(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["A1"] = 2
    sheet["A2"] = 3
    sheet["A3"] = "=SUM(A1:A2)"
    path = _save_workbook(tmp_path, workbook)

    record = _by_cell(build_cell_index(str(path), "file-1", path.name))[("Sheet1", "A3")]

    assert record["formula"] == "=SUM(A1:A2)"
    assert record["value"] in (None, "")


def test_build_cell_index_empty_file_does_not_crash(tmp_path):
    workbook = Workbook()
    path = _save_workbook(tmp_path, workbook)

    assert build_cell_index(str(path), "file-1", path.name) == []


def test_build_cell_index_labels_headers_and_row_context(tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["C1"] = "Unit Price"
    sheet["A2"] = "Line 1"
    sheet["B2"] = "Product"
    sheet["C2"] = 1.83
    path = _save_workbook(tmp_path, workbook)

    record = _by_cell(build_cell_index(str(path), "file-1", path.name))[("Sheet1", "C2")]

    assert record["left_label"] == "Product"
    assert record["above_header"] == "Unit Price"
    assert record["row_context"] == "Line 1|Product|1.83"


def test_build_cell_index_returns_empty_for_unreadable_workbook(tmp_path):
    path = tmp_path / "broken.xlsx"
    path.write_bytes(b"not-an-xlsx")

    assert build_cell_index(str(path), "file-1", path.name) == []
