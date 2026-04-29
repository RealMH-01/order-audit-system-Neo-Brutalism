import io
import json
import sys
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from openpyxl import Workbook
from openpyxl import load_workbook

from app.services.report_generator import ReportGeneratorService


def _save_workbook(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet["F9"] = "HR-EXP250401"
    workbook.save(path)
    return path


def _xlsx_issue() -> dict:
    return {
        "id": "R-01",
        "level": "RED",
        "field_name": "发票号 / 合同号",
        "finding": "发票号与合同号不一致",
        "suggestion": "请核对原始单据",
        "confidence": 0.98,
        "file_id": "xlsx-1",
        "mark_status": "marked",
        "mark_reason_code": "MARKED",
        "locations": [
            {
                "file_id": "xlsx-1",
                "file_name": "CI-HR-EXP2504001.xlsx",
                "sheet": "Sheet1",
                "cell": "F9",
                "confidence": 0.98,
                "resolver": "anchor_verified",
            }
        ],
    }


def _filenames() -> dict[str, str]:
    return {
        "marked": "审核标记版-HR-EXP250401-20260429-1555.xlsx",
        "detailed": "审核详情版-HR-EXP250401-20260429-1555.xlsx",
        "zip": "审核报告-HR-EXP250401-20260429-1555.zip",
    }


def test_zip_uses_final_structure_and_writes_task_info_and_manifest(tmp_path):
    source = _save_workbook(tmp_path / "CI-HR-EXP2504001.xlsx")
    context = {
        "uploaded_files": [
            {
                "id": "xlsx-1",
                "filename": "CI-HR-EXP2504001.xlsx",
                "extension": "xlsx",
                "original_xlsx_path": str(source),
            }
        ],
        "original_xlsx_paths": {"xlsx-1": str(source)},
    }
    audit_result = {
        "confidence": 0.91,
        "summary": {"red": 1, "yellow": 0, "blue": 0},
        "issues": [_xlsx_issue()],
    }

    report_zip = ReportGeneratorService().generate_report_zip(
        "task-zip-1",
        audit_result,
        context,
        _filenames(),
    )

    with zipfile.ZipFile(report_zip) as zip_file:
        namelist = zip_file.namelist()
        assert namelist == [
            "审核详情版-HR-EXP250401-20260429-1555.xlsx",
            "标记版/审核标记版-CI-HR-EXP2504001-20260429-1555.xlsx",
            "任务信息.txt",
            "manifest.json",
        ]
        task_info_bytes = zip_file.read("任务信息.txt")
        assert task_info_bytes[:3] == b"\xef\xbb\xbf"
        task_info_text = task_info_bytes.decode("utf-8-sig")
        assert "task-zip-1" in task_info_text
        assert "业务标识" in task_info_text
        assert "本次上传文件清单" in task_info_text
        assert "审核结果摘要" in task_info_text

        manifest = json.loads(zip_file.read("manifest.json").decode("utf-8"))
        assert manifest["schema_version"] == "1.0"
        assert manifest["task_id"] == "task-zip-1"
        assert manifest["summary"]["total_issues"] == 1
        assert manifest["files"][0]["marked_version_generated"] is True
        assert "finding" not in manifest["issues"][0]
        assert "suggestion" not in manifest["issues"][0]


def test_report_bundle_uses_marked_only_zip_and_keeps_full_zip_structure(tmp_path):
    source = _save_workbook(tmp_path / "CI-HR-EXP2504001.xlsx")
    context = {
        "uploaded_files": [
            {
                "id": "xlsx-1",
                "filename": "CI-HR-EXP2504001.xlsx",
                "extension": "xlsx",
                "original_xlsx_path": str(source),
            }
        ],
        "original_xlsx_paths": {"xlsx-1": str(source)},
    }
    audit_result = {
        "confidence": 0.91,
        "summary": {"red": 1, "yellow": 0, "blue": 0},
        "issues": [_xlsx_issue()],
    }

    bundle = ReportGeneratorService().generate_report_bundle("task-bundle-1", audit_result, context)

    assert bundle["filenames"]["marked"].endswith(".zip")
    with zipfile.ZipFile(bundle["marked_report"]) as marked_zip:
        marked_names = marked_zip.namelist()
        assert len(marked_names) == 1
        assert marked_names[0].startswith("标记版/审核标记版-")
        assert marked_names[0].endswith(".xlsx")
        assert "任务信息.txt" not in marked_names
        assert "manifest.json" not in marked_names

    with zipfile.ZipFile(bundle["report_zip"]) as full_zip:
        full_names = full_zip.namelist()
        assert any(name.startswith("审核详情版-") and name.endswith(".xlsx") for name in full_names)
        assert any(name.startswith("标记版/审核标记版-") and name.endswith(".xlsx") for name in full_names)
        assert "任务信息.txt" in full_names
        assert "manifest.json" in full_names
        detailed_name = next(name for name in full_names if name.startswith("审核详情版-"))
        detailed_workbook = load_workbook(io.BytesIO(full_zip.read(detailed_name)))
        assert detailed_workbook.sheetnames[:2] == ["问题明细", "审核摘要"]


def test_all_pdf_upload_omits_empty_marked_directory(tmp_path):
    context = {
        "uploaded_files": [
            {
                "id": "pdf-1",
                "filename": "报关单.pdf",
                "extension": "pdf",
            }
        ]
    }
    audit_result = {
        "confidence": 0.76,
        "summary": {"red": 0, "yellow": 1, "blue": 0},
        "issues": [
            {
                "id": "Y-01",
                "level": "YELLOW",
                "field_name": "报关资料",
                "file_id": "pdf-1",
                "document_label": "报关单.pdf",
                "mark_status": "unsupported_file_type",
                "mark_reason_code": "FILE_TYPE_NOT_EXCEL",
                "locations": [],
            }
        ],
    }

    report_zip = ReportGeneratorService().generate_report_zip(
        "task-pdf-1",
        audit_result,
        context,
        _filenames(),
    )

    with zipfile.ZipFile(report_zip) as zip_file:
        namelist = zip_file.namelist()
        assert "审核详情版-HR-EXP250401-20260429-1555.xlsx" in namelist
        assert "任务信息.txt" in namelist
        assert "manifest.json" in namelist
        assert not any(name.startswith("标记版/") for name in namelist)
        manifest = json.loads(zip_file.read("manifest.json").decode("utf-8"))
        assert manifest["schema_version"] == "1.0"
        assert manifest["summary"]["by_mark_status"]["unsupported_file_type"] == 1
