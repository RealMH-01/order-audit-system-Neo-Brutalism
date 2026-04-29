"""Build a reliable cell-coordinate index for uploaded Excel workbooks."""

from __future__ import annotations

import logging
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import coordinate_to_tuple, get_column_letter, range_boundaries

logger = logging.getLogger(__name__)


def build_cell_index(
    xlsx_path: str,
    file_id: str,
    file_name: str,
    document_type: str | None = None,
) -> list[dict]:
    """解析 .xlsx，返回非空单元格坐标索引列表"""

    resolved_document_type = document_type or _infer_document_type(file_name)
    try:
        formula_workbook = load_workbook(xlsx_path, data_only=False)
        value_workbook = load_workbook(xlsx_path, data_only=True)
    except Exception:
        logger.warning("Failed to load workbook for cell index: %s", xlsx_path, exc_info=True)
        return []

    index: list[dict] = []
    try:
        for formula_sheet in formula_workbook.worksheets:
            sheet_name = formula_sheet.title
            try:
                if sheet_name not in value_workbook.sheetnames:
                    logger.warning("Skipping sheet missing in data_only workbook: %s", sheet_name)
                    continue
                value_sheet = value_workbook[sheet_name]
                index.extend(
                    _build_sheet_index(
                        formula_sheet=formula_sheet,
                        value_sheet=value_sheet,
                        file_id=file_id,
                        file_name=file_name,
                        document_type=resolved_document_type,
                    )
                )
            except Exception:
                logger.warning("Failed to build cell index for sheet %s in %s", sheet_name, xlsx_path, exc_info=True)
                continue
    finally:
        formula_workbook.close()
        value_workbook.close()

    return index


def normalize_merged_cell(sheet_name: str, cell: str, merged_ranges: list[str]) -> str:
    """
    如果 cell 落在某个 merged_range 内，返回该合并区域左上角主格；
    否则返回原 cell。
    """

    try:
        row, column = coordinate_to_tuple(cell)
    except Exception:
        logger.warning("Invalid cell coordinate for merged normalization: sheet=%s cell=%s", sheet_name, cell)
        return cell

    for merged_range in merged_ranges:
        try:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
        except Exception:
            logger.warning(
                "Invalid merged range for normalization: sheet=%s range=%s",
                sheet_name,
                merged_range,
            )
            continue
        if min_row <= row <= max_row and min_col <= column <= max_col:
            return f"{get_column_letter(min_col)}{min_row}"
    return cell


def _build_sheet_index(
    *,
    formula_sheet: Any,
    value_sheet: Any,
    file_id: str,
    file_name: str,
    document_type: str,
) -> list[dict]:
    merged_ranges = [str(merged_range) for merged_range in formula_sheet.merged_cells.ranges]
    merged_lookup = _build_merged_lookup(merged_ranges)
    value_strings: dict[tuple[int, int], str] = {}
    cell_records: list[dict[str, Any]] = []

    max_row = formula_sheet.max_row or 0
    max_column = formula_sheet.max_column or 0

    for row_idx in range(1, max_row + 1):
        for column_idx in range(1, max_column + 1):
            merged_info = merged_lookup.get((row_idx, column_idx))
            if merged_info and not merged_info["is_anchor"]:
                continue

            formula_cell = formula_sheet.cell(row=row_idx, column=column_idx)
            value_cell = value_sheet.cell(row=row_idx, column=column_idx)
            raw_value = formula_cell.value
            cached_value = value_cell.value
            formula = raw_value if isinstance(raw_value, str) and raw_value.startswith("=") else None
            display_value = cached_value
            if display_value in (None, "") and formula is not None:
                display_value = formula
            value_str = _stringify(display_value)

            if value_str == "" and formula is None:
                continue

            value_strings[(row_idx, column_idx)] = value_str
            cell_records.append(
                {
                    "file_id": file_id,
                    "file_name": file_name,
                    "document_type": document_type,
                    "sheet": formula_sheet.title,
                    "cell": formula_cell.coordinate,
                    "row": row_idx,
                    "column": column_idx,
                    "value": cached_value if cached_value not in ("",) else None,
                    "value_str": value_str,
                    "formula": formula,
                    "number_format": formula_cell.number_format,
                    "merged_range": merged_info["range"] if merged_info else None,
                    "is_merge_anchor": bool(merged_info and merged_info["is_anchor"]),
                    "left_label": None,
                    "above_header": None,
                    "row_context": "",
                }
            )

    row_contexts = _build_row_contexts(value_strings)
    for record in cell_records:
        row = int(record["row"])
        column = int(record["column"])
        record["left_label"] = _find_left_label(value_strings, row, column)
        record["above_header"] = _find_above_header(value_strings, row, column)
        record["row_context"] = row_contexts.get(row, "")

    return cell_records


def _build_merged_lookup(merged_ranges: list[str]) -> dict[tuple[int, int], dict[str, Any]]:
    lookup: dict[tuple[int, int], dict[str, Any]] = {}
    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = range_boundaries(merged_range)
        for row_idx in range(min_row, max_row + 1):
            for column_idx in range(min_col, max_col + 1):
                lookup[(row_idx, column_idx)] = {
                    "range": merged_range,
                    "is_anchor": row_idx == min_row and column_idx == min_col,
                }
    return lookup


def _build_row_contexts(value_strings: dict[tuple[int, int], str]) -> dict[int, str]:
    row_values: dict[int, list[tuple[int, str]]] = {}
    for (row, column), value in value_strings.items():
        if value:
            row_values.setdefault(row, []).append((column, value))
    return {
        row: "|".join(value for _, value in sorted(values, key=lambda item: item[0]))[:200]
        for row, values in row_values.items()
    }


def _find_left_label(value_strings: dict[tuple[int, int], str], row: int, column: int) -> str | None:
    for current_column in range(column - 1, 0, -1):
        value = value_strings.get((row, current_column), "").strip()
        if len(value) >= 1:
            return value
    return None


def _find_above_header(value_strings: dict[tuple[int, int], str], row: int, column: int) -> str | None:
    for current_row in range(row - 1, 0, -1):
        value = value_strings.get((current_row, column), "").strip()
        if len(value) >= 1:
            return value
    return None


def _stringify(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _infer_document_type(file_name: str) -> str:
    stem = Path(file_name).stem
    lowered = stem.lower()
    uppered = stem.upper()
    if "CI" in uppered or "invoice" in lowered:
        return "invoice"
    if "PL" in uppered or "packing" in lowered:
        return "packing_list"
    if "PO" in uppered:
        return "po"
    if "报关" in stem:
        return "customs"
    return "unknown"
