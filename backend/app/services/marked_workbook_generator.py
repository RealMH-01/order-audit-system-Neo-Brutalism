"""Generate marked copies from original uploaded Excel workbooks."""

from __future__ import annotations

import copy
import re
import shutil
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import PatternFill
from openpyxl.utils.cell import coordinate_to_tuple

_LEVEL_COLORS = {
    "RED": "FF6B6B",
    "YELLOW": "FFD93D",
    "BLUE": "6BCBFF",
}
_LEVEL_RANK = {"BLUE": 1, "YELLOW": 2, "RED": 3}
_LEVEL_SORT = {"RED": 0, "YELLOW": 1, "BLUE": 2}
_WINDOWS_ILLEGAL_CHARS = re.compile(r'[<>:"/\\|?*]')
_CONTROL_WHITESPACE = re.compile(r"[\n\r\t]")
_MULTIPLE_UNDERSCORES = re.compile(r"_+")


def merge_levels(levels: list[str]) -> str:
    """RED > YELLOW > BLUE，输入多个返回最高级"""

    normalized = [_normalize_level(level) for level in levels]
    if not normalized:
        return "YELLOW"
    return max(normalized, key=lambda level: _LEVEL_RANK[level])


def generate_marked_copies(
    issues: list[dict],
    uploaded_files: dict,
    output_dir: Path,
    timestamp: str,
) -> tuple[list[Path], list[dict], list[dict]]:
    """
    返回：
    1. 生成成功的标记版文件路径列表；
    2. 文件级生成结果摘要；
    3. 更新后的 issues 副本，用于详情版和 manifest。
    """

    updated_issues = copy.deepcopy(issues)
    file_records = _normalize_uploaded_files(uploaded_files)
    output_dir.mkdir(parents=True, exist_ok=True)

    grouped = _group_marked_locations(updated_issues)
    skipped = _group_skipped_file_issues(updated_issues, file_records, grouped)
    generated_paths: list[Path] = []
    summaries: list[dict] = []

    for file_id, entries in grouped.items():
        record = file_records.get(file_id, {"id": file_id})
        file_name = str(record.get("filename") or _first_location_file_name(entries) or f"{file_id}.xlsx")
        extension = _file_extension(record, file_name)
        summary = {
            "file_id": file_id,
            "file_name": file_name,
            "status": "skipped",
            "reason": "",
            "output_path": "",
            "marked_issue_count": len({entry["issue_index"] for entry in entries}),
            "marked_cell_count": 0,
        }

        if extension != "xlsx":
            summary["reason"] = f"unsupported_file_type:{extension or 'unknown'}"
            summaries.append(summary)
            continue

        source_path = _source_path(record)
        if source_path is None:
            summary["status"] = "failed"
            summary["reason"] = "missing_original_xlsx_path"
            _mark_entries_failed(updated_issues, entries, "missing_original_xlsx_path")
            summaries.append(summary)
            continue

        copy_path = output_dir / _marked_filename(file_name, timestamp)
        try:
            shutil.copy2(source_path, copy_path)
            workbook = load_workbook(copy_path)
            marked_cells = _apply_marks(workbook, entries, updated_issues)
            workbook.save(copy_path)
        except Exception as exc:
            summary["status"] = "failed"
            summary["reason"] = str(exc)
            _mark_entries_failed(updated_issues, entries, str(exc))
            try:
                if copy_path.exists():
                    copy_path.unlink()
            except OSError:
                pass
            summaries.append(summary)
            continue

        summary["status"] = "generated"
        summary["output_path"] = str(copy_path)
        summary["marked_cell_count"] = marked_cells
        summaries.append(summary)
        generated_paths.append(copy_path)

    for file_id, issue_indexes in skipped.items():
        record = file_records.get(file_id, {"id": file_id})
        file_name = str(record.get("filename") or f"{file_id}")
        extension = _file_extension(record, file_name)
        summaries.append(
            {
                "file_id": file_id,
                "file_name": file_name,
                "status": "skipped",
                "reason": f"unsupported_file_type:{extension or 'unknown'}",
                "output_path": "",
                "marked_issue_count": len(issue_indexes),
                "marked_cell_count": 0,
            }
        )

    return generated_paths, summaries, updated_issues


def _normalize_uploaded_files(uploaded_files: dict) -> dict[str, dict[str, Any]]:
    records: dict[str, dict[str, Any]] = {}
    if not isinstance(uploaded_files, dict):
        return records

    original_paths = uploaded_files.get("original_xlsx_paths")
    if isinstance(original_paths, dict):
        for file_id, path in original_paths.items():
            records[str(file_id)] = {"id": str(file_id), "original_xlsx_path": str(path)}

    record_items: list[Any]
    if isinstance(uploaded_files.get("uploaded_files"), list):
        record_items = uploaded_files["uploaded_files"]
    elif isinstance(uploaded_files.get("files"), list):
        record_items = uploaded_files["files"]
    else:
        record_items = [
            {"id": key, "original_xlsx_path": value} if isinstance(value, (str, Path)) else value
            for key, value in uploaded_files.items()
            if key not in {"original_xlsx_paths", "cell_indexes"}
        ]

    for item in record_items:
        if not isinstance(item, dict):
            continue
        file_id = str(item.get("id") or item.get("file_id") or "")
        if not file_id:
            continue
        record = dict(records.get(file_id) or {})
        record.update(item)
        path = record.get("original_xlsx_path") or record.get("path") or record.get("file_path")
        if path:
            record["original_xlsx_path"] = str(path)
        records[file_id] = record

    return records


def _group_marked_locations(issues: list[dict]) -> dict[str, list[dict[str, Any]]]:
    grouped: dict[str, list[dict[str, Any]]] = {}
    for issue_index, issue in enumerate(issues):
        if not isinstance(issue, dict) or issue.get("mark_status") != "marked":
            continue
        locations = issue.get("locations")
        if not isinstance(locations, list):
            continue
        for location in locations:
            if not isinstance(location, dict):
                continue
            file_id = str(location.get("file_id") or issue.get("file_id") or "")
            if not file_id:
                continue
            grouped.setdefault(file_id, []).append(
                {
                    "issue_index": issue_index,
                    "location": location,
                }
            )
    return grouped


def _group_skipped_file_issues(
    issues: list[dict],
    file_records: dict[str, dict[str, Any]],
    marked_groups: dict[str, list[dict[str, Any]]],
) -> dict[str, set[int]]:
    skipped: dict[str, set[int]] = {}
    for issue_index, issue in enumerate(issues):
        if not isinstance(issue, dict) or issue.get("mark_status") == "marked":
            continue
        file_id = _issue_file_id(issue)
        if not file_id or file_id in marked_groups:
            continue
        record = file_records.get(file_id, {"id": file_id})
        file_name = str(record.get("filename") or issue.get("document_label") or file_id)
        extension = _file_extension(record, file_name)
        if extension and extension != "xlsx":
            skipped.setdefault(file_id, set()).add(issue_index)
    return skipped


def _issue_file_id(issue: dict[str, Any]) -> str:
    if issue.get("file_id"):
        return str(issue["file_id"])
    for key in ("locations", "candidate_locations"):
        locations = issue.get(key)
        if not isinstance(locations, list):
            continue
        for location in locations:
            if isinstance(location, dict) and location.get("file_id"):
                return str(location["file_id"])
    return ""


def _apply_marks(workbook: Any, entries: list[dict[str, Any]], issues: list[dict]) -> int:
    cell_issue_map: dict[tuple[str, str], dict[int, dict[str, Any]]] = {}

    for entry in entries:
        issue_index = int(entry["issue_index"])
        location = entry["location"]
        sheet_name = str(location.get("sheet") or "").strip()
        cell_ref = str(location.get("cell") or "").strip()
        if not sheet_name or not cell_ref or sheet_name not in workbook.sheetnames:
            continue
        worksheet = workbook[sheet_name]
        try:
            coordinate_to_tuple(cell_ref)
        except ValueError:
            continue
        anchor_cell = _merged_anchor_cell(worksheet, cell_ref)
        key = (sheet_name, anchor_cell)
        cell_issue_map.setdefault(key, {})[issue_index] = location

    for (sheet_name, cell_ref), indexed_locations in cell_issue_map.items():
        worksheet = workbook[sheet_name]
        cell = worksheet[cell_ref]
        issue_items = [issues[index] for index in indexed_locations]
        level = merge_levels([str(issue.get("level", "YELLOW")) for issue in issue_items])
        cell.fill = PatternFill(fill_type="solid", fgColor=f"FF{_LEVEL_COLORS[level]}")
        cell.comment = Comment(_comment_text(issue_items), "审核系统")

    return len(cell_issue_map)


def _merged_anchor_cell(worksheet: Any, cell_ref: str) -> str:
    for merged_range in worksheet.merged_cells.ranges:
        if cell_ref in merged_range:
            return merged_range.start_cell.coordinate
    return cell_ref


def _comment_text(issues: list[dict]) -> str:
    ordered = sorted(
        issues,
        key=lambda issue: (_LEVEL_SORT[_normalize_level(str(issue.get("level", "YELLOW")))], str(issue.get("id", ""))),
    )
    lines: list[str] = []
    if len(ordered) > 1:
        lines.append(f"本单元格命中 {len(ordered)} 个问题")

    for issue in ordered:
        level = _normalize_level(str(issue.get("level", "YELLOW")))
        finding = str(issue.get("finding") or issue.get("message") or "").strip()
        suggestion = str(issue.get("suggestion") or "").strip()
        lines.append(f"[{level}] {finding}")
        lines.append(f"建议: {suggestion}")

    return "\n".join(lines)


def _mark_entries_failed(issues: list[dict], entries: list[dict[str, Any]], reason: str) -> None:
    for issue_index in {int(entry["issue_index"]) for entry in entries}:
        issue = issues[issue_index]
        issue["mark_status"] = "mark_failed"
        issue["mark_reason_code"] = "WRITE_FAILED"
        issue["mark_reason"] = reason


def _marked_filename(file_name: str, timestamp: str) -> str:
    stem = _sanitize_identifier(Path(file_name).stem)
    safe_timestamp = _sanitize_identifier(timestamp)
    return f"审核标记版-{stem}-{safe_timestamp}.xlsx"


def _sanitize_identifier(identifier: str) -> str:
    text = _WINDOWS_ILLEGAL_CHARS.sub("_", str(identifier or ""))
    text = _CONTROL_WHITESPACE.sub("", text)
    text = text.strip(" .")
    text = _MULTIPLE_UNDERSCORES.sub("_", text)
    text = text[:80]
    return text or "未命名"


def _normalize_level(level: str) -> str:
    normalized = str(level or "").upper().strip()
    return normalized if normalized in _LEVEL_RANK else "YELLOW"


def _source_path(record: dict[str, Any]) -> Path | None:
    raw_path = record.get("original_xlsx_path") or record.get("path") or record.get("file_path")
    if not raw_path:
        return None
    path = Path(str(raw_path))
    return path if path.exists() else None


def _file_extension(record: dict[str, Any], file_name: str) -> str:
    extension = str(record.get("extension") or "").strip().lower().lstrip(".")
    if extension:
        return extension
    return Path(file_name).suffix.lower().lstrip(".")


def _first_location_file_name(entries: list[dict[str, Any]]) -> str:
    for entry in entries:
        location = entry.get("location")
        if isinstance(location, dict) and location.get("file_name"):
            return str(location["file_name"])
    return ""
