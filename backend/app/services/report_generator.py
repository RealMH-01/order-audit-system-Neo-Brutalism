"""报告生成服务：输出标记版、详情版 Excel，并支持 ZIP 打包。"""

from __future__ import annotations

import io
import json
import re
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any

from app.errors import AppError
from app.models.schemas import FeatureStatus
from app.services.marked_workbook_generator import generate_marked_copies
from app.services.report_filename import build_report_filename, pick_report_identifier
from app.services.report_manifest import build_manifest
from app.services.task_info_writer import render_task_info_text

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Font, PatternFill
except Exception:  # pragma: no cover
    Workbook = None
    Comment = None
    Font = None
    PatternFill = None

_LEVEL_FILLCOLORS = {
    "RED": "FECACA",
    "YELLOW": "FEF08A",
    "BLUE": "BFDBFE",
}

_MARKED_REPORT_MESSAGES = {
    "NO_MARKABLE_XLSX": "本次任务没有可生成原表标记副本的 Excel 文件。",
    "ONLY_UNSUPPORTED_FILES": "本次上传文件暂不支持生成原表标记副本。",
    "NO_MARKED_WORKBOOK": "本次审核没有成功生成可下载的原表标记副本。",
    "LEGACY_TASK_NO_MARKED_WORKBOOK": "该历史任务未保留可下载的原表标记副本，请重新运行审核后下载。",
    "MARKED_REPORT_UNAVAILABLE": "标记版报告暂不可用，请重新运行审核后下载。",
}


class ReportGeneratorService:
    """根据审核结果生成内存态 Excel 与 ZIP 报告。"""

    def get_features(self) -> list[FeatureStatus]:
        """返回当前报告生成模块能力说明。"""

        return [
            FeatureStatus(
                name="标记版与详情版报告导出",
                ready=True,
                note="已支持 Excel 输出、单元格标色、批注置信度和 ZIP 打包。",
            )
        ]

    def generate_marked_report(self, task_id: str, audit_result: dict[str, Any]) -> io.BytesIO:
        """
        Deprecated, no longer used by user-facing endpoints. Kept temporarily to avoid cascading test rewrites.
        Do NOT call from new code.
        """

        workbook = self._new_workbook()
        sheet = workbook.active
        sheet.title = "审核问题标记汇总"
        sheet.append(["任务 ID", task_id])
        sheet.append(["红色", audit_result.get("summary", {}).get("red", 0)])
        sheet.append(["黄色", audit_result.get("summary", {}).get("yellow", 0)])
        sheet.append(["蓝色", audit_result.get("summary", {}).get("blue", 0)])
        sheet.append([])

        headers = ["序号", "级别", "字段", "问题说明", "建议", "文档类型", "文件 ID"]
        sheet.append(headers)
        self._set_column_widths(sheet, [12, 16, 20, 50, 40, 16, 20])
        self._style_header_row(sheet, row_index=6)

        for index, issue in enumerate(self._extract_issues(audit_result), start=1):
            level = issue.get("level", "YELLOW")
            row = [
                issue.get("id", f"issue-{index:03d}"),
                self._localize_level(str(level)),
                issue.get("field_name", "unspecified_field"),
                issue.get("finding") or issue.get("message", ""),
                issue.get("suggestion", ""),
                issue.get("document_type", ""),
                issue.get("file_id", ""),
            ]
            sheet.append(row)
            current_row = sheet.max_row
            self._mark_issue_row(sheet, current_row, str(level), float(issue.get("confidence", 0.5)))

        self._apply_wrap_text(sheet, start_row=1)
        sheet.freeze_panes = "A7"
        return self._workbook_to_bytes(workbook)

    def generate_marked_only_zip(
        self,
        task_id: str,
        audit_result: dict[str, Any],
        audit_context: Any | None = None,
    ) -> io.BytesIO:
        """Generate a ZIP containing only source workbook marked copies."""

        filenames = self.build_report_filenames(task_id, audit_context)
        generated_at = self._generated_at()
        timestamp = self._timestamp_from_filenames(filenames) or self._filename_timestamp(generated_at)

        with tempfile.TemporaryDirectory(prefix="audit-marked-") as temp_dir:
            marked_paths, _, _ = self._generate_marked_artifacts(
                audit_result,
                audit_context,
                Path(temp_dir),
                timestamp,
            )
            if not marked_paths:
                self._raise_marked_report_unavailable(audit_context)
            return self._build_marked_only_zip(marked_paths)

    def generate_detail_report(self, task_id: str, audit_result: dict[str, Any]) -> io.BytesIO:
        """生成详情版 Excel 报告。"""

        workbook = self._new_workbook()
        issue_sheet = workbook.active
        issue_sheet.title = "问题明细"
        headers = [
            "序号",
            "级别",
            "字段",
            "问题说明",
            "建议",
            "原文件名",
            "原表位置",
            "定位置信度",
            "标记状态",
            "未标记原因",
            "文档类型",
            "文件 ID",
            "置信度",
            "PO 值",
            "观察值",
            "引用片段",
        ]
        issue_sheet.append(headers)
        self._set_column_widths(issue_sheet, [12, 16, 20, 50, 40, 24, 28, 14, 18, 36, 16, 20, 12, 24, 24, 40])
        self._style_header_row(issue_sheet, row_index=1)

        for index, issue in enumerate(self._extract_issues(audit_result), start=1):
            level = issue.get("level", "YELLOW")
            issue_sheet.append(
                [
                    issue.get("id", f"issue-{index:03d}"),
                    self._localize_level(str(level)),
                    issue.get("field_name", "unspecified_field"),
                    issue.get("finding") or issue.get("message", ""),
                    issue.get("suggestion", ""),
                    self._source_file_name(issue),
                    self._location_refs(issue),
                    self._location_confidence(issue),
                    issue.get("mark_status", ""),
                    "" if issue.get("mark_status") == "marked" else issue.get("mark_reason", ""),
                    issue.get("document_type", ""),
                    issue.get("file_id", ""),
                    f"{float(issue.get('confidence', 0.5)) * 100:.1f}%",
                    issue.get("matched_po_value") or issue.get("source_value", ""),
                    issue.get("observed_value") or issue.get("your_value", ""),
                    issue.get("source_excerpt", ""),
                ]
            )
            current_row = issue_sheet.max_row
            self._mark_issue_row(issue_sheet, current_row, str(level), float(issue.get("confidence", 0.5)))

        self._apply_wrap_text(issue_sheet, start_row=1)
        issue_sheet.freeze_panes = "A2"

        summary_sheet = workbook.create_sheet("审核摘要")
        summary_sheet.append(["任务 ID", task_id])
        summary_sheet.append(["整体置信度", f"{float(audit_result.get('confidence', 0.5)) * 100:.1f}%"])
        summary_sheet.append(["红色问题数", audit_result.get("summary", {}).get("red", 0)])
        summary_sheet.append(["黄色问题数", audit_result.get("summary", {}).get("yellow", 0)])
        summary_sheet.append(["蓝色问题数", audit_result.get("summary", {}).get("blue", 0)])
        summary_sheet.append(["备注", "\n".join(audit_result.get("notes", [])) or ""])
        return self._workbook_to_bytes(workbook)

    def generate_report_zip(
        self,
        task_id: str,
        audit_result: dict[str, Any],
        audit_context: Any | None = None,
        filenames: dict[str, str] | None = None,
    ) -> io.BytesIO:
        """把面向用户的标记版和详情版报告打包为 ZIP。"""

        filenames = filenames or self.build_report_filenames(task_id, audit_context)
        generated_at = self._generated_at()
        timestamp = self._timestamp_from_filenames(filenames) or self._filename_timestamp(generated_at)
        identifier = pick_report_identifier(audit_context or {}, task_id)

        with tempfile.TemporaryDirectory(prefix="audit-marked-") as temp_dir:
            marked_paths, marked_summary, updated_issues = self._generate_marked_artifacts(
                audit_result,
                audit_context,
                Path(temp_dir),
                timestamp,
            )
            updated_result = {**audit_result, "issues": updated_issues}
            detailed = self.generate_detail_report(task_id, updated_result)
            return self._build_zip_archive(
                task_id=task_id,
                identifier=identifier,
                generated_at=generated_at,
                detailed=detailed,
                detailed_filename=filenames["detailed"],
                marked_paths=marked_paths,
                marked_summary=marked_summary,
                uploaded_files=self._uploaded_files(audit_context),
                issues=updated_issues,
                confidence=audit_result.get("confidence"),
            )

    def generate_report_bundle(
        self,
        task_id: str,
        audit_result: dict[str, Any],
        audit_context: Any | None = None,
    ) -> dict[str, Any]:
        """返回一组可直接下载或持久化的报告二进制对象。"""

        filenames = self.build_report_filenames(task_id, audit_context)
        generated_at = self._generated_at()
        timestamp = self._timestamp_from_filenames(filenames) or self._filename_timestamp(generated_at)
        identifier = pick_report_identifier(audit_context or {}, task_id)

        with tempfile.TemporaryDirectory(prefix="audit-marked-") as temp_dir:
            marked_paths, marked_summary, updated_issues = self._generate_marked_artifacts(
                audit_result,
                audit_context,
                Path(temp_dir),
                timestamp,
            )
            updated_result = {**audit_result, "issues": updated_issues}
            detailed_report = self.generate_detail_report(task_id, updated_result)
            report_zip = self._build_zip_archive(
                task_id=task_id,
                identifier=identifier,
                generated_at=generated_at,
                detailed=detailed_report,
                detailed_filename=filenames["detailed"],
                marked_paths=marked_paths,
                marked_summary=marked_summary,
                uploaded_files=self._uploaded_files(audit_context),
                issues=updated_issues,
                confidence=audit_result.get("confidence"),
            )
            marked_report = self._build_marked_only_zip(marked_paths) if marked_paths else None

        bundle = {
            "detailed_report": detailed_report,
            "report_zip": report_zip,
            "filenames": filenames,
        }
        if marked_report is not None:
            bundle["marked_report"] = marked_report
        return bundle

    @staticmethod
    def build_report_filenames(task_id: str, audit_context: Any | None = None) -> dict[str, str]:
        """Build user-facing names for all generated report artifacts."""

        identifier = pick_report_identifier(audit_context or {}, task_id)
        marked_filename = build_report_filename("标记版", identifier, "zip")
        timestamp = ReportGeneratorService._timestamp_from_filenames({"marked": marked_filename})
        return {
            "marked": marked_filename,
            "detailed": ReportGeneratorService._with_filename_timestamp(
                build_report_filename("详情版", identifier, "xlsx"),
                timestamp,
            ),
            "zip": ReportGeneratorService._with_filename_timestamp(
                build_report_filename("报告", identifier, "zip"),
                timestamp,
            ),
        }

    def _generate_marked_artifacts(
        self,
        audit_result: dict[str, Any],
        audit_context: Any | None,
        output_dir: Path,
        timestamp: str,
    ) -> tuple[list[Path], list[dict], list[dict]]:
        """Generate source-workbook marked copies for ZIP packaging."""

        return generate_marked_copies(
            self._extract_issues(audit_result),
            self._context_mapping(audit_context),
            output_dir,
            timestamp,
        )

    def _build_zip_archive(
        self,
        *,
        task_id: str,
        identifier: str,
        generated_at: str,
        detailed: io.BytesIO,
        detailed_filename: str,
        marked_paths: list[Path],
        marked_summary: list[dict],
        uploaded_files: list[dict],
        issues: list[dict],
        confidence: Any,
    ) -> io.BytesIO:
        archive = io.BytesIO()
        summary_payload = {"files": marked_summary, "confidence": confidence}
        task_info = render_task_info_text(
            task_id=task_id,
            identifier=identifier,
            generated_at=generated_at,
            uploaded_files=uploaded_files,
            issues=issues,
            marked_summary=summary_payload,
        )
        manifest = build_manifest(
            task_id=task_id,
            identifier=identifier,
            generated_at=generated_at,
            uploaded_files=uploaded_files,
            issues=issues,
            marked_files=marked_summary,
        )

        with zipfile.ZipFile(archive, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(detailed_filename, detailed.getvalue())
            for marked_path in marked_paths:
                zf.write(marked_path, arcname=f"标记版/{marked_path.name}")
            zf.writestr("任务信息.txt", task_info.encode("utf-8-sig"))
            zf.writestr("manifest.json", json.dumps(manifest, ensure_ascii=False, indent=2).encode("utf-8"))

        archive.seek(0)
        return archive

    @staticmethod
    def _build_marked_only_zip(marked_paths: list[Path]) -> io.BytesIO:
        archive = io.BytesIO()
        written = 0
        with zipfile.ZipFile(archive, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
            for marked_path in marked_paths:
                if marked_path.suffix.lower() != ".xlsx":
                    continue
                zf.write(marked_path, arcname=f"标记版/{marked_path.name}")
                written += 1
        if written == 0:
            raise AppError(
                _MARKED_REPORT_MESSAGES["MARKED_REPORT_UNAVAILABLE"],
                status_code=409,
                reason_code="MARKED_REPORT_UNAVAILABLE",
            )
        archive.seek(0)
        return archive

    @classmethod
    def _raise_marked_report_unavailable(cls, audit_context: Any | None) -> None:
        reason_code = cls._marked_report_reason_code(audit_context)
        raise AppError(_MARKED_REPORT_MESSAGES[reason_code], status_code=409, reason_code=reason_code)

    @classmethod
    def _marked_report_reason_code(cls, audit_context: Any | None) -> str:
        context = cls._context_mapping(audit_context)
        uploaded_files = cls._uploaded_files(context)
        original_paths = context.get("original_xlsx_paths")
        has_original_xlsx = isinstance(original_paths, dict) and any(str(path or "").strip() for path in original_paths.values())
        has_uploaded_xlsx = any(cls._file_extension(item) == "xlsx" for item in uploaded_files)
        if has_original_xlsx or has_uploaded_xlsx:
            return "NO_MARKED_WORKBOOK"
        if uploaded_files and all(cls._file_extension(item) != "xlsx" for item in uploaded_files):
            return "ONLY_UNSUPPORTED_FILES"
        return "NO_MARKABLE_XLSX"

    @staticmethod
    def _file_extension(file_record: dict[str, Any]) -> str:
        extension = str(file_record.get("extension") or "").strip().lower().lstrip(".")
        if extension:
            return extension
        filename = str(file_record.get("filename") or file_record.get("name") or "")
        return Path(filename).suffix.lower().lstrip(".")

    @staticmethod
    def _context_mapping(audit_context: Any | None) -> dict[str, Any]:
        if isinstance(audit_context, dict):
            return audit_context
        if audit_context is None:
            return {}
        return {
            key: getattr(audit_context, key)
            for key in ("uploaded_files", "files", "original_xlsx_paths", "cell_indexes")
            if hasattr(audit_context, key)
        }

    @classmethod
    def _uploaded_files(cls, audit_context: Any | None) -> list[dict]:
        context = cls._context_mapping(audit_context)
        for key in ("uploaded_files", "files", "file_records"):
            value = context.get(key)
            if isinstance(value, list):
                return [dict(item) for item in value if isinstance(item, dict)]

        files: list[dict] = []
        original_paths = context.get("original_xlsx_paths")
        if isinstance(original_paths, dict):
            for file_id, path in original_paths.items():
                files.append(
                    {
                        "id": str(file_id),
                        "filename": Path(str(path)).name,
                        "extension": "xlsx",
                        "original_xlsx_path": str(path),
                    }
                )
        return files

    @staticmethod
    def _generated_at() -> str:
        return datetime.now().astimezone().replace(microsecond=0).isoformat()

    @staticmethod
    def _filename_timestamp(generated_at: str) -> str:
        try:
            generated = datetime.fromisoformat(generated_at)
        except ValueError:
            generated = datetime.now().astimezone()
        return generated.strftime("%Y%m%d-%H%M")

    @staticmethod
    def _timestamp_from_filenames(filenames: dict[str, str]) -> str:
        for key in ("zip", "detailed", "marked"):
            match = re.search(r"-(\d{8}-\d{4})\.[^.]+$", str(filenames.get(key) or ""))
            if match:
                return match.group(1)
        return ""

    @staticmethod
    def _with_filename_timestamp(filename: str, timestamp: str) -> str:
        if not timestamp:
            return filename
        return re.sub(r"-\d{8}-\d{4}(\.[^.]+)$", rf"-{timestamp}\1", filename)

    @staticmethod
    def _source_file_name(issue: dict[str, Any]) -> str:
        locations = issue.get("locations")
        if isinstance(locations, list) and locations:
            first = locations[0]
            if isinstance(first, dict) and first.get("file_name"):
                return str(first["file_name"])
        candidates = issue.get("candidate_locations")
        if isinstance(candidates, list) and candidates:
            first = candidates[0]
            if isinstance(first, dict) and first.get("file_name"):
                return str(first["file_name"])
        return str(issue.get("document_label") or "")

    @staticmethod
    def _location_refs(issue: dict[str, Any]) -> str:
        locations = issue.get("locations")
        if not locations and issue.get("mark_status") == "multiple_candidates":
            locations = issue.get("candidate_locations")
        if not isinstance(locations, list):
            return ""

        refs: list[str] = []
        for location in locations:
            if not isinstance(location, dict):
                continue
            sheet = str(location.get("sheet") or "").strip()
            cell = str(location.get("cell") or "").strip()
            if sheet and cell:
                refs.append(f"{sheet}!{cell}")
        return "; ".join(dict.fromkeys(refs))

    @staticmethod
    def _location_confidence(issue: dict[str, Any]) -> str:
        locations = issue.get("locations")
        if not locations and issue.get("mark_status") == "low_confidence":
            locations = issue.get("candidate_locations")
        if not isinstance(locations, list) or not locations:
            return ""

        values = [
            float(location["confidence"])
            for location in locations
            if isinstance(location, dict) and location.get("confidence") not in (None, "")
        ]
        if not values:
            return ""
        return "; ".join(dict.fromkeys(f"{value:.2f}" for value in values))

    def _new_workbook(self):
        """创建 Workbook 并检查依赖。"""

        if Workbook is None or PatternFill is None or Comment is None or Font is None:
            raise AppError("当前环境缺少 openpyxl，暂时无法生成报告。", status_code=500)
        return Workbook()

    @staticmethod
    def _extract_issues(audit_result: dict[str, Any]) -> list[dict[str, Any]]:
        """从审核结果中提取问题列表。"""

        issues = audit_result.get("issues")
        if isinstance(issues, list):
            return [item for item in issues if isinstance(item, dict)]
        return []

    @staticmethod
    def _style_header_row(sheet, row_index: int) -> None:
        """设置表头样式。"""

        for cell in sheet[row_index]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(fill_type="solid", fgColor="D9EAF7")

    @staticmethod
    def _mark_issue_row(sheet, row_index: int, level: str, confidence: float) -> None:
        """对问题行做精准标色，并在说明列写入置信度批注。"""

        fill_color = _LEVEL_FILLCOLORS.get(level.upper(), _LEVEL_FILLCOLORS["YELLOW"])
        fill = PatternFill(fill_type="solid", fgColor=fill_color)
        for column in ("B", "C", "D", "E"):
            sheet[f"{column}{row_index}"].fill = fill
        sheet[f"D{row_index}"].comment = Comment(
            f"模型置信度：{max(0.0, min(1.0, confidence)) * 100:.1f}%",
            "system",
        )

    @staticmethod
    def _localize_level(level: str) -> str:
        """将英文级别标签转换为中文显示标签。"""

        mapping = {"RED": "红色·高风险", "YELLOW": "黄色·疑点", "BLUE": "蓝色·提示"}
        return mapping.get(level.upper().strip(), level)

    @staticmethod
    def _set_column_widths(sheet, widths: list[int]) -> None:
        """批量设置列宽。"""

        from openpyxl.utils import get_column_letter

        for index, width in enumerate(widths, start=1):
            sheet.column_dimensions[get_column_letter(index)].width = width

    @staticmethod
    def _apply_wrap_text(sheet, start_row: int) -> None:
        """对指定起始行之后的所有数据行设置自动换行。"""

        from openpyxl.styles import Alignment

        for row in sheet.iter_rows(min_row=start_row, max_row=sheet.max_row):
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    @staticmethod
    def _workbook_to_bytes(workbook) -> io.BytesIO:
        """把 Workbook 输出成 BytesIO。"""

        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)
        return buffer
